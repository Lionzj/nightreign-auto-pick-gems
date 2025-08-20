

import pandas as pd
from itertools import combinations
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import time
import os
 
script_path = os.path.dirname(os.path.abspath(__file__))

# 全局常量 - 关键词条
REQUIRED_PHRASES = {
    "每次打倒封印监牢里的囚犯，能永久性提升攻击力",
    # "受到损伤的当下，能通过攻击恢复部分血量",
    # "受到损伤并被弹飞时，能提升强韧度与减伤率",
    # "【执行者】在绝招发动期间，能以咆哮恢复血量",
    # "出击时，武器战技改为“冻霜踏地”",
    "提升专注值上限",
    "连续攻击时，恢复专注值",
    "出击时的武器，附加异常状态出血",
    "出击时，武器战技改为“血刃”"
    # "出击时，会持有“石剑钥匙”"
}

# 允许的颜色组合（列表中的每个子列表代表一种有效组合）
ALLOWED_COLOR_COMBOS = [
        ['红', '黄', '黄'],  # 需要1红、2黄
        ['红', '蓝', '绿'],  # 需要1红、1蓝、1绿
        ['红', '红', '蓝'],  # 需要2红、1蓝
        ['黄', '黄', '红'],  # 需要2黄、1红
        ['黄', '黄', '黄'],  # 需要3黄
        ['绿', '绿', '绿'],  # 需要3绿
        ['蓝', '蓝', '蓝'],  # 需要3蓝
        ['蓝', '黄']    
]

def timer_decorator(func):
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        print(f"{func.__name__} executed in {end-start:.2f} seconds")
        return result
    return wrapper

def check_color_combo(actual_colors, required_combo):
    """
    严格检查颜色组合
    actual_colors: 实际颜色列表，如 ['红','黄','绿']
    required_combo: 要求的颜色组合，如 ['红','黄','黄']
    返回: bool
    """
    actual_count = Counter(actual_colors)
    required_count = Counter(required_combo)
    
    # 检查每种必需颜色的最小数量
    for color, count in required_count.items():
        if actual_count[color] < count:
            return False
    
    # 特殊处理蓝黄组合（允许第三色任意）
    if set(required_combo) == {'蓝', '黄'}:
        return True
    
    # 其他组合必须完全匹配颜色数量
    return sum(actual_count.values()) == sum(required_count.values())

@timer_decorator
def load_and_preprocess(file_path):
    df = pd.read_excel(file_path, sheet_name='宝石库', usecols='A:E')
    # 确保词列为字符串类型
    for col in ['词条一', '词条二', '词条三']:
        df[col] = df[col].astype(str)
    return df

@timer_decorator
def find_valid_combinations(df):
    valid_combos = []
    
    # 预筛选包含至少一个关键词条的宝石
    filtered_rows = [
        row for _, row in df.iterrows()
        if any(any(req in phrase for req in REQUIRED_PHRASES)
               for phrase in [row['词条一'], row['词条二'], row['词条三']])
    ]
    
    # 检查所有可能的3宝石组合
    for combo in combinations(filtered_rows, 3):
        colors = [row['颜色'] for row in combo]
        
        # 检查颜色组合
        color_ok = False
        for req_combo in ALLOWED_COLOR_COMBOS:
            if check_color_combo(colors, req_combo):
                color_ok = True
                break
        if not color_ok:
            continue
        
        # 检查词条覆盖
        all_phrases = '|'.join(
            f"{row['词条一']}|{row['词条二']}|{row['词条三']}" 
            for row in combo
        )
        covered = set()
        for req in REQUIRED_PHRASES:
            if req in all_phrases:
                covered.add(req)
        
        if covered == REQUIRED_PHRASES:
            valid_combos.append(combo)
    
    return valid_combos

@timer_decorator
def export_to_excel(combinations, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "宝石组合"
    
    # 设置样式
    red_font = Font(color="FF0000")
    header_fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    
    # 写入表头
    headers = ["组合ID", "宝石位置", "序号", "颜色", "词条一", "词条二", "词条三"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    # 写入数据
    combo_id = 1
    for combo in combinations:
        for i, row in enumerate(combo, 1):
            ws.append([
                combo_id,
                f"宝石{i}",
                row['序号'],
                row['颜色'],
                row['词条一'],
                row['词条二'],
                row['词条三']
            ])
            
            # 标记关键词条
            for col in range(5, 8):  # 词条列
                cell = ws.cell(row=ws.max_row, column=col)
                if any(req in str(cell.value) for req in REQUIRED_PHRASES):
                    cell.font = red_font
        
        # 添加空行分隔
        ws.append([])
        combo_id += 1
    
    # 设置列宽
    col_widths = {'A': 8, 'B': 10, 'C': 8, 'D': 8, 'E': 65, 'F': 65, 'G': 65}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output_path)

if __name__ == "__main__":
    input_path = f'{script_path}/宝石库.xlsx'
    output_path = f'{script_path}/执行者.xlsx'

    print("开始处理...")
    df = load_and_preprocess(input_path)
    valid_combos = find_valid_combinations(df)
    export_to_excel(valid_combos, output_path)
    
    print(f"\n结果已保存至: {output_path}")
    print(f"共找到 {len(valid_combos)} 个符合条件的组合")
